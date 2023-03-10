from math import floor
from  openpyxl import load_workbook, utils
from random import choice
from time import time,sleep
import tkinter as tk
from threading import Thread

name_file = 'exel_document.xlsx'
# Обработчик ошибок
try: workbook = load_workbook(name_file)
except FileNotFoundError: print(f"Файл {name_file} не найден."); input(); exit()
except utils.exceptions.InvalidFileException: print("Файл некорректен или поврежден."); input(); exit()

# tkinter
win = tk.Tk()  # Переменная создания главного окна
h = 450  # Высота окна
w = 400  # Ширина окна
win.config(bg='#F4A900')  # Цвет основного заднего фона окна приложения
win.title('Learner foreign words')  # Название окна приложения
win.geometry(f"{h}x{w}")  # Указание размеров окна (см. 19 и 20 строки)
win.resizable(False, False)  # Указание того что размеры окна нельзя будет изменить

# openpyxl
sheet = workbook['Sheet1'] # Выбираем нужный лист
words_counter = sheet.max_row # Кол-во строк в документе
words = {} # Хранение слов

# извлечение слов
remove_xa0 = lambda s: s.replace('\xa0', '') # удаление пробелом в кодировке ASCII
for i in range(1, words_counter + 1):
    eng_word = remove_xa0(sheet.cell(row=i, column=1).value).lower() # чтение первого столбца
    rus_word = remove_xa0(sheet.cell(row=i, column=2).value).lower() # чтение второго столбца
    words[eng_word] = rus_word # запись обеих значений в словарь

# Проверка на правильность ответа
# def check_value(answer, random_value):
#     if answer in random_value.lower() and len(answer) != 0: return True
#     else: return False

# ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓   КОНСОЛЬНОЕ ПРИЛОЖЕНИЕ   ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓

# def Check_Knowleague_Hunan(words): # Консольный ввод/вывод
#     start_time = time() # старт секундомера
#     last_value, random_value, fail_counter = '','',0
#     random_words = words # клон словаря со значениями в котором после успешного ответа удаляется слово
#     while len(random_words) > 0: # Пока в клоне словаря со значениями есть значения(по мере успешных ответов их кол-во уменьшается)
#         last_value = random_value # запоминание прошлого правильного ответа для подсказки
#         random_key = choice(list(words.keys())) # берёт случайный ключ
#         random_value = words[random_key] # по случайному ключу берём значение
#         answer = input(f'{random_key} - ') # получаем ответ пользователя
#         if check_value(answer, random_value): # если ответ верный удаляем данное слово
#             print('Верно!')
#             del random_words[random_key]
#         elif answer == '1':
#             print(f'Ответ: {random_value}')
#         elif answer == '2':
#             print(f'Ответ на прошлый вопрос: {last_value}')
#         else:
#             print('Не верно((')
#             fail_counter+=1
#         print(f'Слов осталось: {len(random_words)}') # счётчик остатка слов
#     elapsed_time = time() - start_time # подсчёт затраченного времени
#     print(f'Поздавляю! Вы успешно ответили на все слова!! На это вам понадобилось {floor(elapsed_time)} секунд')
#     print(f'Число неудачных ответов: {fail_counter}')

# print("Введите 1 для текущего ответа и 2 для прошлого ответа")
# print("\033[1;32m\033[40mВведите 1 для текущего ответа и 2 для прошлого ответа\033[0m") # вывод красивого сообщения которое работает только в IDE
# Check_Knowleague_Hunan(words) # работа консоли

# ↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑   КОНСОЛЬНОЕ ПРИЛОЖЕНИЕ   ↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


# ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓   ГРАФИЧЕСКОЕ ПРИЛОЖЕНИЕ   ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓

def check_value(answer, random_value):
    if answer.lower() in random_value.lower() and len(answer) != 0: return True
    else: return False

# лютый колхоз, желаю удачи если тебе здесь что-то надо :)
def Check_Knowleague_Hunan_GUI(words):
    start_time = time()  # старт секундомера
    last_value, random_value, fail_counter, tips_counter = '', '', 0, 0 # прошлый ответ, текущий правильный ответ, счётчик ошибок и счётчик подсказок
    random_words = words  # клон словаря со значениями в котором после успешного ответа удаляется слово
    while len(random_words) > 0:  # Пока в клоне словаря со значениями есть значения(по мере успешных ответов их кол-во уменьшается)
        words_count.configure(text=f'Осталось слов: {len(random_words)}') # меняем значение tk.label просле каждого ответа
        fails_count.configure(text=f'Неудачных ответов: {fail_counter}') # меняем значение tk.label после каждого неудачного ответа
        if Done_Button.cget("state") != "active" and Check_Last_Answer_Button.cget("state") != "active" and Check_Answer_Button.cget("state") != "active": # баг, проверяем что бы ни одна из кнопок не была нажата
            last_value = random_value  # запоминание прошлого правильного ответа для подсказки
            random_key = choice(list(words.keys()))  # берёт случайный ключ
            random_value = words[random_key]  # по случайному ключу берём значение
            Random_word.configure(text=random_key) # заменяем tk.label на случайное слово
            while Done_Button.cget("state") != "active" and Check_Last_Answer_Button.cget("state") != "active" and Check_Answer_Button.cget("state") != "active": # ещё один баг, проверяем что бы ни одна из кнопок не была нажата, иначе программа не работает так как задумано
                sleep(0.005) # что бы пк не взорвался
                time_count.configure(text=f'Прошло уже {floor(time() - start_time)} секунд') # обновляем значение tk.label после каждой секунды
                if Done_Button.cget("state") == "active": # Если нажата кнопка "проверить"
                    answer = InputAnswer.get("1.0", tk.END) # присваиваем значение tk.label переменной (то что написал пользователь)
                    InputAnswer.delete("1.0", "end") # очищаем tk.label, что бы пользователю не пришлось делать это вручную
                    if check_value(answer.strip(), random_value): # если ответ правильный
                        WindowOutput.insert(tk.END, f"{random_key}- {answer}", "green") # выводим зелёный текст
                        del random_words[random_key] # удаляем данное слово
                    else: # иначе (то есть если ответ не верный)
                        WindowOutput.insert(tk.END, f"{random_key}- {answer}", "red") # выводим красный текст
                        fail_counter+=1 # увеличиваем счётчик неудачных ответов
                elif Check_Last_Answer_Button.cget("state") == "active": # если кнопка "узнать прошлый ответ" нажата
                    WindowOutput.insert(tk.END, f"Прошлый ответ- {last_value}\n", "blue") # вывести синий текст
                    tips_counter+=1 # счётчик использованных подсказок +1
                    tips_count.configure(text=f'Подсказок: {tips_counter}') # обновляем вывод кол-во использованых подсказок
                    break # покидаем данный while и идём к самому первому while в этой функции
                elif Check_Answer_Button.cget("state") == "active": # здёсь всё тоже самое что и на несколько стрчоек выше, но только для узнавания текущего ответа
                    WindowOutput.insert(tk.END, f"Текущий ответ- {random_value}\n", "blue")
                    tips_counter+=1
                    tips_count.configure(text=f'Подсказок: {tips_counter}')
                    break
    # обновление текста после окончания всех слов
    words_count.configure(text=f'Слова кончились!')
    Random_word.configure(text='')
    time_count.configure(text=f'Прошло {floor(time() - start_time)} секунд')



# Первая строка с текстами
Text_0 = tk.Label(win, text='История:', bg='#F4A900', font = ('Arial', 14))
Name_file = tk.Label(win, text=f'{name_file}', bg='#F4A900', font = ('Arial', 14))

# Окно c выводом основного текста
WindowOutput = tk.Text(height=9, bg="#FADFAD", fg='black', font = ('Helvetica', 16))

# Ввод пользователем данных
Random_word = tk.Label(win, text='Случайное слово:', bg='#F4A900', font = ('Arial', 14))
InputAnswer = tk.Text(height=1, width=15, bg="#FADFAD", fg='black', font = ('Helvetica', 16))
Done_Button = tk.Button(win, text="Проверить", bg="#FADFAD",activebackground='#F4C430')
Check_Last_Answer_Button = tk.Button(win, text="Узнать прошлый правильный ответ", bg="#FADFAD",activebackground='#F4C430')
Check_Answer_Button = tk.Button(win, text="Узнать правильный ответ", bg="#FADFAD",activebackground='#F4C430')

# вывод доп. инфы
words_count = tk.Label(win, text='Осталось слов:', bg='#F4A900', font = ('Arial', 14))
fails_count = tk.Label(win, text='Неудачных ответов:', bg='#F4A900', font = ('Arial', 14))
time_count = tk.Label(win, text='Прошло уже много секунд', bg='#F4A900', font = ('Arial', 14))
tips_count = tk.Label(win, text='Подсказок: 0', bg='#F4A900', font = ('Arial', 14))



# Первая строка с текстами
Text_0.place(anchor="nw")
Name_file.place(anchor="nw",relx=1.0, x=-len(name_file)*10)

# Окно c выводом основного текста и добавление конфингов для цветного текста
WindowOutput.pack(pady=30,fill="x", padx=7)
WindowOutput.tag_config("red", foreground="red")
WindowOutput.tag_config("green", foreground="green")
WindowOutput.tag_config("blue", foreground="#0099cc")

# Ввод пользователем данных
Random_word.place(y=260)
InputAnswer.place(y=260, x=170)
Done_Button.place(y=263, x=360)
Check_Last_Answer_Button.place(y=295, x=223)
Check_Answer_Button.place(y=327, x=280)
words_count.place(y=295)
fails_count.place(y=327)
time_count.place(y=359)
tips_count.place(y=359, x=280)

# создание потока для фоновой работы и обновления данных GUI
thread_0 = Thread(target=Check_Knowleague_Hunan_GUI, daemon=True, args=(words,))
thread_0.start()

win.mainloop()  # работа в граф. окне

# ↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑   ГРАФИЧЕСКОЕ ПРИЛОЖЕНИЕ   ↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑